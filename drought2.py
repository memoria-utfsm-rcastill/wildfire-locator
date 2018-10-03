import pymongo
import openpyxl
import argparse
import json

from datetime import datetime


class Rain:
    mm = 0
    ts = None

    def __init__(self, mm, ts):
        self.mm = mm
        self.ts = ts


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-m', '--mongodb', default='localhost',
                        help='mongo DB connection string, default: localhost')
    parser.add_argument('-d', '--db', default='dmc',
                        help='mongo DB Database, default: dmc')
    parser.add_argument('-c', '--coll', default='data',
                        help='mongo DB Collection, default: data')
    parser.add_argument('-o', '--output', required=True,
                        help='file to write output (extension required, either xlsx|json)')
    parser.add_argument('result',
                        help='file of wildfire occurrences on area (json file, result of locator.py)')

    args = parser.parse_args()

    if not args.output.endswith('.json') and not args.output.endswith('.xlsx'):
        print('Output type must be either xlsx|json')

    args.mongodb = 'mongodb://{uri}'.format(
        uri=args.mongodb.replace('mongodb://', ''))

    mclient = pymongo.MongoClient(args.mongodb)
    coll = mclient[args.db][args.coll]

    with open(args.result) as f:
        wildfire = json.load(f)

    for entry in wildfire:
        entry['start'] = datetime.strptime(entry['start'], '%Y-%m-%d %H:%M:%S')

    # Sort from earliest to latest
    wildfire.sort(key=lambda x: x['start'])

    # Consider only after 3 wildfires (Julio's method)
    wildfire = [(wildfire[i-2], entry)
                for i, entry in enumerate(wildfire) if (i + 1) % 3 == 0]

    # Load rain periods
    rain_periods = []
    current_rain = None
    for doc in coll.find({}):
        # if valid
        if doc['ts'] != doc['prc_ts']:
            # if is rain
            if doc['prc'] != 0:
                # add to current (contiguous)
                if current_rain != None:
                    current_rain.mm += doc['prc']
                    current_rain.ts = doc['ts']
                # new rain
                else:
                    current_rain = Rain(doc['prc'], doc['ts'])
            # end of period
            elif current_rain != None:
                rain_periods.append(current_rain)
                current_rain = None
    # last rain lasted more than data
    if current_rain != None:
        rain_periods.append(current_rain)
    # Periods are sorted from latest to earliest
    rain_periods.sort(key=lambda x: x.ts, reverse=True)

    time_series = []
    for first, third in wildfire:
        rain_before_fire = next(
            (rain for rain in rain_periods if rain.ts < first['start']), None)
        if rain_before_fire == None:
            print('WRN: There is no rain before {ts}'.format(first['start']))
            continue
        time_series.append({'mm': rain_before_fire.mm, 'diff': (
            third['start']-rain_before_fire.ts).days})

    '''legacy
    time_series = []
    for period in rain_periods:
        third = next(
            (t['start'] for f, t in wildfire if f['start'] > period.ts), None)
        if third == None:
            print('No wildfires after rain: {r}'.format(r=period.ts))
            break
        time_series.append({'mm': period.mm, 'diff': (third-period.ts).days})
    '''

    if args.output.endswith('xlsx'):
        wb = openpyxl.Workbook()

        del wb[wb.sheetnames[0]]

        ws = wb.create_sheet('Drought Data')

        ws.cell(row=1, column=1, value='Rain (mm)')
        ws.cell(row=1, column=2, value='Days Passed')

        for i, entry in enumerate(time_series):
            ws.cell(row=i+2, column=1, value=entry['mm'])
            ws.cell(row=i+2, column=2, value=entry['diff'])

        wb.save(args.output)
        wb.close()
    else:
        # json
        with open(args.output, 'w') as f:
            json.dump(time_series, f)


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print()
