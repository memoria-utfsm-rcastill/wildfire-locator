import pymongo
import openpyxl
import argparse
import json

from datetime import datetime


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-m', '--mongodb', default='localhost',
                        help='mongo DB connection string, default: localhost')
    parser.add_argument('-d', '--db', default='dmc',
                        help='mongo DB Database, default: dmc')
    parser.add_argument('-c', '--coll', default='data',
                        help='mongo DB Collection, default: data')
    parser.add_argument('-b', '--beginning-of-times', default='2013-01-01T00:00:00',
                        help='reference timestamp to calculate days passed')
    parser.add_argument('-o', '--output', required=True,
                        help='file to write output (extension required, either xlsx|json)')
    parser.add_argument('result',
                        help='file of wildfire occurrences on area (json file, result of locator.py)')

    args = parser.parse_args()

    if not args.output.endswith('.json') and not args.output.endswith('.xlsx'):
        print('Output type must be either xlsx|json')

    args.mongodb = 'mongodb://{uri}'.format(
        uri=args.mongodb.replace('mongodb://', ''))
    args.beginning_of_times = datetime.strptime(
        args.beginning_of_times, '%Y-%m-%dT%H:%M:%S')

    mclient = pymongo.MongoClient(args.mongodb)
    coll = mclient[args.db][args.coll]

    with open(args.result) as f:
        wildfire = json.load(f)

    for entry in wildfire:
        entry['start'] = datetime.strptime(entry['start'], '%Y-%m-%d %H:%M:%S')

    # Sort from earliest to latest
    wildfire.sort(key=lambda x: x['start'])

    # Consider only after 3 wildfires (Julio's method)
    wildfire = [entry for i, entry in enumerate(wildfire) if (i + 1) % 3 == 0]

    time_series = []

    last_occurrence = args.beginning_of_times
    for fire in wildfire:
        diff = fire['start'] - last_occurrence

        # diff.days <- this
        cursor = coll.find(
            {'ts': {'$gte': last_occurrence, '$lte': fire['start']}})
        real_acc = 0
        pred_acc = 0
        for entry in cursor:
            if entry['ts'] == entry['prc_ts']:
                real_acc += entry['prc']
            pred_acc += entry['prc']

        time_series.append(
            {'realAcc': real_acc, 'predAcc': pred_acc, 'daysPassed': diff.days})

        last_occurrence = fire['start']

    if args.output.endswith('xlsx'):
        wb = openpyxl.Workbook()

        del wb[wb.sheetnames[0]]

        ws = wb.create_sheet('Drought Data')

        ws.cell(row=1, column=1, value='Real Acc')
        ws.cell(row=1, column=2, value='Pred Acc')
        ws.cell(row=1, column=3, value='Days Passed')

        for i, entry in enumerate(time_series):
            ws.cell(row=i+2, column=1, value=entry['realAcc'])
            ws.cell(row=i+2, column=2, value=entry['predAcc'])
            ws.cell(row=i+2, column=3, value=entry['daysPassed'])

        wb.save(args.output)
        wb.close()
    else:
        # json
        with open(args.output, 'w') as f:
            json.dump(time_series, f)


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print()
