import openpyxl
import argparse
import re
import json

from datetime import datetime

COORD = re.compile(
    '(?P<deg>\d{1,3})°(?P<min>\d{1,2})\'(?P<sec>\d{1,2})" *(?P<sgn>[NSOE])')

SGN = {
    'N': 1,
    'S': -1,
    'O': -1,
    'E': 1
}

MON = {
    'ene': '01',
    'feb': '02',
    'mar': '03',
    'abr': '04',
    'may': '05',
    'jun': '06',
    'jul': '07',
    'ago': '08',
    'sep': '09',
    'oct': '10',
    'nov': '11',
    'dic': '12'
}


class WildfireEntry:
    lat = None
    lon = None
    start = None

    @staticmethod
    def from_dict(d):
        return WildfireEntry(d['lat'], d['lon'], d['start'], from_dict=True)

    def __init__(self, lat, lon, start, *, from_dict=False, row=-1):
        if from_dict:
            self.lat = lat
            self.lon = lon
            self.start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
            return
        try:
            lat_match = COORD.match(lat)
            if lat_match is None:
                raise ValueError('Invalid latitude: {l}'.format(l=lat))
            lon_match = COORD.match(lon)
            if lon_match is None:
                raise ValueError('Invalid longitude: {l}'.format(l=lon))
        except Exception:
            print('lat={},lon={},strt={},row={}'.format(lat, lon, start, row))
            raise
        self.lat = SGN[lat_match.group('sgn')] * (float(lat_match.group('deg')) + float(lat_match.group('min')) /
                                                  60 + float(lat_match.group('sec')) / 3600)
        self.lon = SGN[lon_match.group('sgn')] * (float(lon_match.group('deg')) + float(lon_match.group('min')) /
                                                  60 + float(lon_match.group('sec')) / 3600)
        if type(start) is datetime:
            self.start = start
        else:
            day, mon, yhm = start.split('-')
            self.start = datetime.strptime('{d}-{m}-{other}'.format(d=day.zfill(2),
                                                                    m=MON[mon], other=yhm), '%d-%m-%Y %H:%M')

    def __repr__(self):
        return '[{dt}] {lat}, {lon}'.format(dt=self.start, lat=self.lat, lon=self.lon)

    def to_dict(self):
        return {'lat': self.lat, 'lon': self.lon, 'start': self.start}


def load_xlsx(args):
    try:
        wb = openpyxl.load_workbook(args.dbfile)
    except FileNotFoundError:
        print('Error: File "{xls}" not found'.format(xls=args.dbfile))
        return

    ws = wb[args.sheet]

    if args.take <= 0:
        args.take = ws.max_row
    taken = 0

    entries = []

    while taken <= args.take:
        # rows start at 1
        taken += 1
        r = taken+args.skip
        lat = ws.cell(row=r, column=args.lat_col).value
        if lat is None or lat == '':
            if args.empty_row_policy == 'ignore':
                continue
            print('Lat is empty at row: {row}'.format(row=r))
            if args.empty_row_policy == 'alert':
                continue
            else:
                # policy: halt
                break
        lon = ws.cell(row=r, column=args.lon_col).value
        if lon is None or lon == '':
            if args.empty_row_policy == 'ignore':
                continue
            print('Lon is empty at row: {row}'.format(row=r))
            if args.empty_row_policy == 'alert':
                continue
            else:
                break
        start = ws.cell(row=r, column=args.start_col).value
        if start is None or start == '':
            if args.empty_row_policy == 'ignore':
                continue
            print('Start is empty at row: {row}'.format(row=r))
            if args.empty_row_policy == 'alert':
                continue
            else:
                break

        try:
            entries.append(WildfireEntry(lat, lon, start, row=r))
        except ValueError as err:
            print(err)
            return

    return entries


def load_json(args):
    try:
        dbfile = open(args.dbfile)
    except FileNotFoundError:
        print('Error: File "{json}" not found'.format(json=args.dbfile))
    with dbfile:
        return [WildfireEntry.from_dict(entry) for entry in json.load(dbfile)]


def quad(dataset, nwlat, nwlon, selat, selon):
    return json.dumps(sorted([entry.to_dict() for entry in dataset if entry.lon >= nwlon and entry.lon <= selon and entry.lat <= nwlat and entry.lat >= selon], key=lambda x: x['start']), default=lambda x: str(x))


def main():
    parser = argparse.ArgumentParser(description='')
    parser.add_argument('dbfile', help='wildfire log (excel or json file)')
    parser.add_argument('--sheet', default='Consolidado temporadas ',
                        help='custom sheet name, default: "Consolidado temporadas "')
    parser.add_argument('--lat-col', type=int, default=9,
                        help='custom latitude column, default: 9')
    parser.add_argument('--lon-col', type=int, default=10,
                        help='custom longitude column, default: 10')
    parser.add_argument('--start-col', type=int, default=29,
                        help='custom "wildfire start date" column, default: 29')
    parser.add_argument('--skip', type=int, default=1,
                        help='custom row skip, default: 1 (header)')
    parser.add_argument('--take', type=int, default=0,
                        help='custom number of rows to consider, default: 0 (all)')
    parser.add_argument('--export', type=str, default='',
                        help='Export xlsx data to refined json file')
    parser.add_argument('--import', dest='load_json', action='store_true',
                        help='use dbfile as .json exported file (--export)')
    parser.add_argument('-p', '--empty-row-policy',
                        choices=['ignore', 'halt', 'alert'], default='halt', help='should the program ignore -or- halt if row is missing data, default: halt')
    parser.add_argument('-q', '--query', type=str, default='',
                        help='query over data, if not specified, program exits after loading data')

    args = parser.parse_args()

    if args.load_json:
        entries = load_json(args)
    else:
        entries = load_xlsx(args)

    if entries == None:
        return

    if args.export != '':
        with open(args.export, 'w') as f:
            # only unsupported value is datetime, str(x) is enough
            json.dump([entry.to_dict()
                       for entry in entries], f, default=lambda x: str(x))

    if args.query == '':
        return

    instructions = args.query.split(';')
    for ins in instructions:
        try:
            typ, data = ins.split('=')
        except ValueError:
            print('Invalid query instruction: {q}'.format(q=ins))
            return

        if typ.lower() == 'quad':
            try:
                nwlat, nwlon, selat, selon = [
                    float(corner) for corner in data.split(',')]
                print(quad(entries, nwlat, nwlon, selat, selon))
            except ValueError:
                print('Invalid quad definition: {d}'.format(d=data))
                return
        else:
            print('Invalid instruction type: {t}'.format(t=typ))
            return


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print()
